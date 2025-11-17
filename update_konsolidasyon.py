"""
Konsolidasyon Raporu Güncelleme Programı

Bu program data.xlsx'teki verileri alarak Konsolidasyon raporunu günceller:
1. data.xlsx'teki 19 kontrol order'dan 17 tanesini işler (18-19 konsolidasyonda yok)
2. %40'lık verileri x2.5 ile çarparak %100'e dönüştürür (5 şirket için)
3. Gerçekleşen Aylık Euro sayfasını günceller (tüm aylar, tüm metrikler için)
4. Finansal Raporlama AY sayfasındaki formülleri kontrol eder
"""

import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import sys
from pathlib import Path
import shutil

# Dosya yolları
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_FILE = BASE_DIR / 'data.xlsx'
DEFAULT_KONSOLIDASYON_FILE = BASE_DIR / 'Konsolidasyon_2025_NV (1).xlsx'

# Şirket bilgileri (data.xlsx ve konsolidasyon'da aynı sıralama)
# Her kontrol order için şirketler aynı pozisyonlarda
# Offset 0 = İnci Holding, 1 = ISM, 2 = İncitaş, 3 = İnci GS Yuasa, ...
COMPANIES = [
    {"name": "İnci Holding", "is_40_percent": False},      # %100'lük - DOKUNMA
    {"name": "ISM", "is_40_percent": False},               # %100'lük - DOKUNMA
    {"name": "İncitaş", "is_40_percent": False},           # %100'lük - DOKUNMA
    {"name": "İnci GS Yuasa", "is_40_percent": True},      # %40'lık - x2.5 YAP
    {"name": "Yusen İnci Lojistik", "is_40_percent": True},# %40'lık - x2.5 YAP
    {"name": "Maxion Jantas", "is_40_percent": True},      # %40'lık - x2.5 YAP
    {"name": "Maxion Celik", "is_40_percent": True},       # %40'lık - x2.5 YAP
    {"name": "Maxion Aluminyum", "is_40_percent": True},   # %40'lık - x2.5 YAP
]

# Kontrol Order Eşleştirmeleri
# Her kontrol order için data.xlsx ve konsolidasyon satır başlangıçları
KONTROL_ORDER_MAPPING = {
    1: {
        "name": "Net Satışlar",
        "data_row_start": 5,      # data.xlsx'de şirketler 5-12
        "kons_row_start": 5,       # konsolidasyon'da şirketler 5-12
    },
    2: {
        "name": "Satılan Malların Maliyeti (SMM)",
        "data_row_start": 15,
        "kons_row_start": 16,
    },
    3: {
        "name": "Brüt Kar",
        "data_row_start": 25,
        "kons_row_start": 27,
    },
    4: {
        "name": "Faaliyet Giderleri",
        "data_row_start": 35,
        "kons_row_start": 38,
    },
    5: {
        "name": "Diğer Faaliyet Giderleri",
        "data_row_start": 45,
        "kons_row_start": 49,
    },
    6: {
        "name": "EBIT",
        "data_row_start": 55,
        "kons_row_start": 60,
    },
    7: {
        "name": "Amortisman",
        "data_row_start": 65,
        "kons_row_start": 71,
    },
    8: {
        "name": "EBITDA",
        "data_row_start": 75,
        "kons_row_start": 81,
    },
    9: {
        "name": "Finansman Gelir-Gideri",
        "data_row_start": 85,
        "kons_row_start": 92,
    },
    10: {
        "name": "Diğer Gelir-Gideri",
        "data_row_start": 95,
        "kons_row_start": 103,
    },
    11: {
        "name": "Vergi Öncesi Kar",
        "data_row_start": 105,
        "kons_row_start": 114,
    },
    12: {
        "name": "Vergi",
        "data_row_start": 115,
        "kons_row_start": 125,
    },
    13: {
        "name": "Net Kar",
        "data_row_start": 125,
        "kons_row_start": 136,
    },
    14: {
        "name": "Krediler",
        "data_row_start": 135,
        "kons_row_start": 147,
    },
    15: {
        "name": "Nakit ve Bankalar",
        "data_row_start": 145,
        "kons_row_start": 160,
    },
    16: {
        "name": "Net Finansal Borç",
        "data_row_start": 155,
        "kons_row_start": 173,
    },
    17: {
        "name": "Yatırımlar",
        "data_row_start": 165,
        "kons_row_start": 186,
    },
    # NOT: Kontrol Order 18 (Operasyonel Nakit) ve 19 (Serbest Nakit)
    # konsolidasyon dosyasında bulunmuyor, atlanıyor.
}

# Ay eşleştirmeleri
# data.xlsx Gerçekleşen sütunları -> Konsolidasyon sütunları
MONTH_MAPPING = {
    5: (2, "Ocak"),      # data col 5 (Gerçekleşen) -> konsolidasyon col B (2)
    7: (3, "Şubat"),
    9: (4, "Mart"),
    11: (5, "Nisan"),
    13: (6, "Mayıs"),
    15: (7, "Haziran"),
    17: (8, "Temmuz"),
    19: (9, "Ağustos"),
    21: (10, "Eylül"),
    23: (11, "Ekim"),
    25: (12, "Kasım"),
    27: (13, "Aralık"),
}


def find_last_month_with_data(ws):
    """data.xlsx'te son veri olan ayı bulur (Kontrol Order 1'e bakarak)"""
    last_month_col = None
    last_month_name = None

    # Kontrol Order 1 (Net Satışlar) üzerinden son ayı tespit et
    ko1_data_start = KONTROL_ORDER_MAPPING[1]["data_row_start"]

    for data_col, (kons_col, month_name) in MONTH_MAPPING.items():
        has_data = False
        # İlk 3 şirketin herhangi birinde veri var mı kontrol et
        for offset in range(3):
            value = ws.cell(ko1_data_start + offset, data_col).value
            if value and value != 0:
                has_data = True
                break

        if has_data:
            last_month_col = data_col
            last_month_name = month_name
        else:
            break  # İlk boş ay bulundu, daha fazla ilerleme

    return last_month_col, last_month_name


def read_data_from_data_xlsx(data_file):
    """
    data.xlsx'ten TÜM kontrol order'ların verilerini okur ve %100'e dönüştürür.

    Returns:
        all_data: {
            kontrol_order: {
                month_col: {
                    'month_name': str,
                    'data': {kons_row: value, ...}
                }
            }
        }
        last_month_name: str
    """
    print("=" * 80)
    print("DATA.XLSX DOSYASINDAN VERİLER OKUNUYOR...")
    print("=" * 80)

    wb = load_workbook(data_file, data_only=True)
    ws = wb['Export']

    # Son veri olan ayı bul
    last_month_col, last_month_name = find_last_month_with_data(ws)
    print(f"\nSon veri olan ay: {last_month_name} (Sütun {last_month_col})")

    # Her kontrol order için verileri oku
    all_data = {}

    for ko_num, ko_info in KONTROL_ORDER_MAPPING.items():
        ko_name = ko_info["name"]
        data_row_start = ko_info["data_row_start"]
        kons_row_start = ko_info["kons_row_start"]

        print(f"\nKontrol Order {ko_num:2d} ({ko_name}) okunuyor...")

        ko_data = {}

        # Her ay için
        for data_col, (kons_col, month_name) in MONTH_MAPPING.items():
            month_data = {}

            # Her şirket için
            for offset, company_info in enumerate(COMPANIES):
                data_row = data_row_start + offset
                kons_row = kons_row_start + offset
                is_40_percent = company_info["is_40_percent"]

                value = ws.cell(data_row, data_col).value

                # %40'lık veriyi %100'e dönüştür
                if value is not None and value != 0:
                    if is_40_percent:
                        converted_value = value * 2.5  # %40 -> %100
                    else:
                        converted_value = value  # Zaten %100
                    month_data[kons_row] = converted_value
                else:
                    month_data[kons_row] = None

            # Bu ayda veri varsa kaydet
            if any(v is not None for v in month_data.values()):
                ko_data[kons_col] = {
                    'month_name': month_name,
                    'data': month_data
                }

        all_data[ko_num] = ko_data

    wb.close()

    print(f"\n{'='*80}")
    print(f"Toplam {len(KONTROL_ORDER_MAPPING)} kontrol order okundu.")
    print(f"Toplam {len(all_data[1])} ay için veri bulundu.")

    return all_data, last_month_name


def update_gercaylık_euro_sheet(all_data, konsolidasyon_file):
    """Gerçekleşen Aylık Euro sayfasını TÜM kontrol order'lar için günceller"""
    print("\n" + "=" * 80)
    print("GERÇEKLEŞEN AYLIK EURO SAYFASI GÜNCELLENİYOR...")
    print("=" * 80)

    wb = load_workbook(konsolidasyon_file)
    ws = wb['gerç aylık-eur']

    total_updates = 0

    # Her kontrol order için
    for ko_num, ko_info in KONTROL_ORDER_MAPPING.items():
        ko_name = ko_info["name"]
        ko_data = all_data[ko_num]

        print(f"\n--- Kontrol Order {ko_num:2d}: {ko_name} ---")

        ko_updates = 0

        # Her ay için
        for kons_col, month_info in ko_data.items():
            month_name = month_info['month_name']
            month_data = month_info['data']

            # Her şirket için
            for kons_row, value in month_data.items():
                if value is not None:
                    old_value = ws.cell(kons_row, kons_col).value
                    ws.cell(kons_row, kons_col).value = value
                    ko_updates += 1
                    total_updates += 1

        print(f"  ✓ {ko_updates} hücre güncellendi")

    # Dosyayı kaydet
    print(f"\n{'='*80}")
    print(f"TOPLAM {total_updates} hücre güncellendi.")
    print("Dosya kaydediliyor...")
    wb.save(konsolidasyon_file)
    wb.close()
    print("✓ Gerçekleşen Aylık Euro sayfası başarıyla güncellendi!")


def update_finansal_ay_formulas(last_month_name, konsolidasyon_file):
    """Finansal Raporlama AY sayfasındaki formülleri son aya göre günceller"""
    print("\n" + "=" * 80)
    print("FİNANSAL RAPORLAMA AY SAYFASI GÜNCELLENİYOR...")
    print("=" * 80)

    # Ay isimlerini A -Döviz sayfasındaki sütunlara eşle
    month_to_columns = {
        "Ocak": ("B", "C"),       # Budget, Actual
        "Şubat": ("E", "F"),
        "Mart": ("H", "I"),
        "Nisan": ("K", "L"),
        "Mayıs": ("N", "O"),
        "Haziran": ("Q", "R"),
        "Temmuz": ("T", "U"),
        "Ağustos": ("W", "X"),
        "Eylül": ("Z", "AA"),
        "Ekim": ("AC", "AD"),
        "Kasım": ("AF", "AG"),
        "Aralık": ("AI", "AJ"),
    }

    if last_month_name not in month_to_columns:
        print(f"❌ Uyarı: '{last_month_name}' ayı için sütun eşleştirmesi bulunamadı!")
        return

    target_budget_col, target_actual_col = month_to_columns[last_month_name]

    print(f"\nSon veri olan ay: {last_month_name}")
    print(f"Hedef sütunlar: {target_budget_col} (Budget) ve {target_actual_col} (Actual)")

    wb = load_workbook(konsolidasyon_file)
    ws = wb['Finansal Raporlama AY']

    update_count = 0

    # Satır 6-14 arası (Toplam + şirketler), tüm sütunları güncelle
    # D, E = Ciro (Budget, Actual)
    # G, H = Brüt Kar
    # J, K = EBITDA
    # M, N = Net Kar
    target_cols = [4, 5, 7, 8, 10, 11, 13, 14]  # D, E, G, H, J, K, M, N

    print(f"\nFormüller güncelleniyor...")

    for row_idx in range(6, 15):  # Satır 6-14 (Toplam + şirketler)
        company_name = ws.cell(row_idx, 2).value  # B sütunu - şirket adı

        for col_idx in target_cols:
            cell = ws.cell(row_idx, col_idx)

            if cell.data_type == 'f' and cell.value:
                old_formula = str(cell.value)

                # Formüldeki tüm ay sütun referanslarını bul ve değiştir
                new_formula = old_formula

                # Tüm aylardaki Budget ve Actual sütunlarını hedef aya güncelle
                for month, (budget_col, actual_col) in month_to_columns.items():
                    # Budget sütununu güncelle (örn: 'Z' -> 'AC')
                    new_formula = new_formula.replace(f"'{budget_col}", f"'__TEMP_BUDGET__")
                    new_formula = new_formula.replace(f"!{budget_col}", f"!__TEMP_BUDGET__")

                    # Actual sütununu güncelle (örn: 'AA' -> 'AD')
                    new_formula = new_formula.replace(f"'{actual_col}", f"'__TEMP_ACTUAL__")
                    new_formula = new_formula.replace(f"!{actual_col}", f"!__TEMP_ACTUAL__")

                # Geçici işaretçileri hedef sütunlarla değiştir
                new_formula = new_formula.replace("__TEMP_BUDGET__", target_budget_col)
                new_formula = new_formula.replace("__TEMP_ACTUAL__", target_actual_col)

                if new_formula != old_formula:
                    cell.value = new_formula
                    update_count += 1
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    if update_count <= 5:  # İlk 5 güncellemeyi göster
                        print(f"  [{row_idx},{col_letter}] {company_name}: Güncellendi")

    if update_count > 0:
        print(f"\n{'='*80}")
        print(f"Toplam {update_count} formül güncellendi.")
        print("Dosya kaydediliyor...")
        wb.save(konsolidasyon_file)
        print("✓ Finansal Raporlama AY sayfası başarıyla güncellendi!")
    else:
        print(f"\n{'='*80}")
        print("⚠ Hiçbir formül güncellenmedi (zaten güncel olabilir).")

    wb.close()


def run_update(data_file, konsolidasyon_file, output_file=None):
    """Data ve konsolidasyon dosyalarını kullanarak güncelleme işlemini yapar."""
    data_path = Path(data_file)
    konsolidasyon_path = Path(konsolidasyon_file)

    if output_file:
        target_path = Path(output_file)
        shutil.copy2(konsolidasyon_path, target_path)
    else:
        target_path = konsolidasyon_path

    # 1. data.xlsx'ten TÜM kontrol order verilerini oku
    all_data, last_month_name = read_data_from_data_xlsx(data_path)

    # 2. Gerçekleşen Aylık Euro sayfasını TÜM metrikler için güncelle
    update_gercaylık_euro_sheet(all_data, target_path)

    # 3. Finansal Raporlama AY sayfasını güncelle
    update_finansal_ay_formulas(last_month_name, target_path)

    return target_path, last_month_name


def main():
    """Ana program"""
    import argparse

    parser = argparse.ArgumentParser(description="Konsolidasyon raporu güncelleme aracı (17 Kontrol Order)")
    parser.add_argument("--data", dest="data_file", default=DEFAULT_DATA_FILE, help="Data dosyası yolu (varsayılan: data.xlsx)")
    parser.add_argument("--konsolidasyon", dest="konsolidasyon_file", default=DEFAULT_KONSOLIDASYON_FILE, help="Konsolidasyon dosyası yolu")
    parser.add_argument("--output", dest="output_file", help="Güncellenmiş dosyanın kaydedileceği yol (varsayılan: konsolidasyon dosyasının üzerine yazar)")

    args = parser.parse_args()

    data_file = Path(args.data_file)
    konsolidasyon_file = Path(args.konsolidasyon_file)
    output_file = Path(args.output_file) if args.output_file else None

    print("\n" + "=" * 80)
    print("KONSOLİDASYON RAPORU GÜNCELLEME PROGRAMI")
    print(f"17 Kontrol Order Desteği ile")
    print("=" * 80)
    print(f"\nData dosyası: {data_file}")
    print(f"Konsolidasyon dosyası: {konsolidasyon_file}")
    if output_file:
        print(f"Çıktı dosyası: {output_file}")

    # Dosyaların varlığını kontrol et
    if not data_file.exists():
        print(f"\n❌ HATA: {data_file} bulunamadı!")
        sys.exit(1)

    if not konsolidasyon_file.exists():
        print(f"\n❌ HATA: {konsolidasyon_file} bulunamadı!")
        sys.exit(1)

    try:
        target_file, last_month_name = run_update(data_file, konsolidasyon_file, output_file)

        print("\n" + "=" * 80)
        print("✓ GÜNCELLEME BAŞARIYLA TAMAMLANDI!")
        print("=" * 80)
        print(f"\nGüncellenmiş dosya: {target_file}")
        print(f"Son güncellenen ay: {last_month_name}")
        print(f"Güncellenen kontrol order sayısı: {len(KONTROL_ORDER_MAPPING)}")

    except Exception as e:
        print(f"\n❌ HATA: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
